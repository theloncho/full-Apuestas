from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.wallet.models import LedgerEntry, AccountType, Direction
from apps.betting.models import Bet, BetStatus, Event, EventStatus, Selection
from apps.betting.services import LiquidationService
from apps.fraud.models import SuspiciousActivity
from apps.audit.models import AuditLog
from django.contrib.auth import get_user_model

User = get_user_model()


@staff_member_required
def dashboard(request):
    """Dashboard del operador con métricas en vivo."""
    # GGR = Total stakes − Total payouts
    total_stakes = LedgerEntry.objects.filter(
        account_type=AccountType.PENDING_BETS,
        direction=Direction.CREDIT,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    total_payouts = LedgerEntry.objects.filter(
        account_type=AccountType.USER_WALLET,
        direction=Direction.CREDIT,
        description__icontains='Payout',
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    ggr = total_stakes - total_payouts

    # Métricas
    active_bets = Bet.objects.filter(status=BetStatus.ACCEPTED).count()
    total_bets = Bet.objects.count()
    active_users = User.objects.filter(account_status='verificado').count()
    pending_fraud = SuspiciousActivity.objects.filter(reviewed=False).count()

    # Eventos en vivo con exposición
    live_events = Event.objects.filter(status=EventStatus.LIVE)
    exposure_data = []
    for event in live_events:
        for market in event.markets.prefetch_related('selections').all():
            for sel in market.selections.all():
                # Exposure = cuánto pierde la casa si gana esta selección
                potential_payouts = Bet.objects.filter(
                    selection=sel,
                    status=BetStatus.ACCEPTED,
                ).aggregate(
                    exposure=Sum(F('stake') * F('odds_at_placement'))
                )['exposure'] or Decimal('0')
                exposure_data.append({
                    'event': str(event),
                    'selection': sel.name,
                    'exposure': potential_payouts,
                })

    # Últimas alertas
    recent_alerts = SuspiciousActivity.objects.filter(reviewed=False)[:10]

    # Últimas apuestas
    recent_bets = Bet.objects.select_related('user', 'selection__market__event').order_by('-placed_at')[:15]

    context = {
        'ggr': ggr,
        'total_stakes': total_stakes,
        'total_payouts': total_payouts,
        'active_bets': active_bets,
        'total_bets': total_bets,
        'active_users': active_users,
        'pending_fraud': pending_fraud,
        'live_events': live_events,
        'exposure_data': exposure_data,
        'recent_alerts': recent_alerts,
        'recent_bets': recent_bets,
    }
    return render(request, 'dashboard/dashboard.html', context)


@staff_member_required
def export_mincetur_report(request):
    """Exporta reporte mensual en formato Excel (.xlsx) estilizado para MINCETUR."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_mincetur.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte MINCETUR"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border_style = Side(border_style="thin", color="DDDDDD")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    headers = [
        'ID Apuesta', 'Usuario', 'DNI', 'Fecha', 'Monto Apostado',
        'Cuota', 'Payout', 'Estado', 'Evento', 'Mercado',
    ]
    
    ws.append(headers)
    
    # Aplicar estilos a la cabecera
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        
    bets = Bet.objects.select_related(
        'user', 'selection__market__event'
    ).order_by('-placed_at')
    
    for row_idx, bet in enumerate(bets, start=2):
        row = [
            str(bet.id),
            bet.user.username,
            bet.user.dni or 'N/A',
            bet.placed_at.strftime('%Y-%m-%d %H:%M:%S'),
            float(bet.stake),
            float(bet.odds_at_placement),
            float(bet.payout) if bet.payout else 0.0,
            bet.get_status_display(),
            str(bet.selection.market.event),
            bet.selection.market.get_market_type_display(),
        ]
        ws.append(row)
        
        # Estilos de celdas de datos
        for col_num, cell in enumerate(ws[row_idx], 1):
            cell.border = border
            if col_num in [5, 6, 7]:  # Columnas numéricas
                cell.number_format = '#,##0.00'
    
    # Auto-ajustar ancho de columnas
    column_widths = {'A': 36, 'B': 15, 'C': 12, 'D': 20, 'E': 15, 'F': 10, 'G': 15, 'H': 15, 'I': 35, 'J': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response


@staff_member_required
def verify_audit_chain(request):
    """Verifica la integridad de la cadena de auditoría."""
    result = AuditLog.verify_chain_integrity()
    recent_logs = AuditLog.objects.order_by('-created_at')[:20]
    return render(request, 'dashboard/audit_verify.html', {
        'result': result,
        'recent_logs': recent_logs,
    })


@staff_member_required
def liquidate_event_view(request, event_id):
    """Vista para liquidar un evento (formulario admin)."""
    event = get_object_or_404(Event, id=event_id)

    if request.method == 'POST':
        try:
            result_home = int(request.POST.get('result_home', 0))
            result_away = int(request.POST.get('result_away', 0))
            LiquidationService.liquidate_event(event, result_home, result_away)
            messages.success(
                request,
                f'Evento liquidado: {event.home_team} {result_home} - {result_away} {event.away_team}'
            )
        except Exception as e:
            messages.error(request, f'Error al liquidar: {str(e)}')
        return redirect('dashboard:dashboard')

    bets_count = Bet.objects.filter(
        selection__market__event=event,
        status=BetStatus.ACCEPTED,
    ).count()

    return render(request, 'dashboard/liquidate_event.html', {
        'event': event,
        'bets_count': bets_count,
    })


@staff_member_required
def fraud_alerts_view(request):
    """Lista de alertas de fraude."""
    alerts = SuspiciousActivity.objects.all()
    if request.GET.get('pending'):
        alerts = alerts.filter(reviewed=False)
    return render(request, 'dashboard/fraud_alerts.html', {
        'alerts': alerts,
    })
